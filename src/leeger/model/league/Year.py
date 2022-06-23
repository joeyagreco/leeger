import random
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from src.leeger.model.abstract.UniqueId import UniqueId
from src.leeger.model.league.Team import Team
from src.leeger.model.league.Week import Week
from src.leeger.model.stat.YearStatSheet import YearStatSheet


@dataclass(kw_only=True)
class Year(UniqueId):
    yearNumber: int
    teams: list[Team]
    weeks: list[Week]

    def statSheet(self, **kwargs) -> YearStatSheet:
        from src.leeger.calculator.year_calculator.AWALCalculator import AWALCalculator
        from src.leeger.calculator.year_calculator.GameOutcomeCalculator import GameOutcomeCalculator
        from src.leeger.calculator.year_calculator.PlusMinusCalculator import PlusMinusCalculator
        from src.leeger.calculator.year_calculator.PointsScoredCalculator import PointsScoredCalculator
        from src.leeger.calculator.year_calculator.ScoringShareCalculator import ScoringShareCalculator
        from src.leeger.calculator.year_calculator.ScoringStandardDeviationCalculator import \
            ScoringStandardDeviationCalculator
        from src.leeger.calculator.year_calculator.SingleScoreCalculator import SingleScoreCalculator
        from src.leeger.calculator.year_calculator.SmartWinsCalculator import SmartWinsCalculator
        from src.leeger.calculator.year_calculator.SSLCalculator import SSLCalculator
        from src.leeger.calculator.year_calculator.YearOutcomeCalculator import YearOutcomeCalculator

        # Game Outcome
        wins = GameOutcomeCalculator.getWins(self, **kwargs)
        losses = GameOutcomeCalculator.getLosses(self, **kwargs)
        ties = GameOutcomeCalculator.getTies(self, **kwargs)
        winPercentage = GameOutcomeCalculator.getWinPercentage(self, **kwargs)
        wal = GameOutcomeCalculator.getWAL(self, **kwargs)
        walPerGame = GameOutcomeCalculator.getWALPerGame(self, **kwargs)

        # AWAL
        awal = AWALCalculator.getAWAL(self, **kwargs)
        awalPerGame = AWALCalculator.getAWALPerGame(self, **kwargs)
        opponentAWAL = AWALCalculator.getOpponentAWAL(self, **kwargs)
        opponentAWALPerGame = AWALCalculator.getOpponentAWALPerGame(self, **kwargs)

        # Smart Wins
        smartWins = SmartWinsCalculator.getSmartWins(self, **kwargs)
        smartWinsPerGame = SmartWinsCalculator.getSmartWinsPerGame(self, **kwargs)
        opponentSmartWins = SmartWinsCalculator.getOpponentSmartWins(self, **kwargs)
        opponentSmartWinsPerGame = SmartWinsCalculator.getOpponentSmartWinsPerGame(self, **kwargs)

        # Points Scored
        pointsScored = PointsScoredCalculator.getPointsScored(self, **kwargs)
        pointsScoredPerGame = PointsScoredCalculator.getPointsScoredPerGame(self, **kwargs)
        opponentPointsScored = PointsScoredCalculator.getOpponentPointsScored(self, **kwargs)
        opponentPointsScoredPerGame = PointsScoredCalculator.getOpponentPointsScoredPerGame(self, **kwargs)

        # Scoring Share
        scoringShare = ScoringShareCalculator.getScoringShare(self, **kwargs)
        opponentScoringShare = ScoringShareCalculator.getOpponentScoringShare(self, **kwargs)

        # Single Score
        maxScore = SingleScoreCalculator.getMaxScore(self, **kwargs)
        minScore = SingleScoreCalculator.getMinScore(self, **kwargs)

        # Scoring Standard Deviation
        scoringStandardDeviation = ScoringStandardDeviationCalculator.getScoringStandardDeviation(self, **kwargs)

        # Plus Minus
        plusMinus = PlusMinusCalculator.getPlusMinus(self, **kwargs)

        # SSL
        teamScore = SSLCalculator.getTeamScore(self, **kwargs)
        teamSuccess = SSLCalculator.getTeamSuccess(self, **kwargs)
        teamLuck = SSLCalculator.getTeamLuck(self, **kwargs)

        # Year Outcome
        championshipCount = YearOutcomeCalculator.getChampionshipCount(self, **kwargs)

        return YearStatSheet(wins=wins, losses=losses, ties=ties, winPercentage=winPercentage, wal=wal,
                             walPerGame=walPerGame, awal=awal, awalPerGame=awalPerGame, opponentAWAL=opponentAWAL,
                             opponentAWALPerGame=opponentAWALPerGame, smartWins=smartWins,
                             smartWinsPerGame=smartWinsPerGame, opponentSmartWins=opponentSmartWins,
                             opponentSmartWinsPerGame=opponentSmartWinsPerGame, pointsScored=pointsScored,
                             pointsScoredPerGame=pointsScoredPerGame, opponentPointsScored=opponentPointsScored,
                             opponentPointsScoredPerGame=opponentPointsScoredPerGame, scoringShare=scoringShare,
                             opponentScoringShare=opponentScoringShare, maxScore=maxScore, minScore=minScore,
                             scoringStandardDeviation=scoringStandardDeviation, plusMinus=plusMinus,
                             teamScore=teamScore, teamSuccess=teamSuccess, teamLuck=teamLuck,
                             championshipCount=championshipCount)

    def toExcel(self, filePath: str, **kwargs) -> None:
        """
        Saves *this* Year to an Excel sheet at the given file path.
        """
        # create workbook and sheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = str(self.yearNumber)

        ###################
        # Styles for table #
        ###################

        # fonts
        headerColumnFont = Font(size=12, bold=True)
        teamNameFont = Font(size=11, bold=True)

        # colors
        GRAY = Color(rgb="B8B8B8")

        def getRandomColor(tint: float = 0) -> Color:
            r = lambda: random.randint(0, 255)
            hexCode = "%02X%02X%02X" % (r(), r(), r())
            return Color(rgb=hexCode, tint=tint)

        TEAM_ROW_COLORS = [getRandomColor(0.5) for _ in range(len(self.teams))]

        # fills
        headerFill = PatternFill(patternType="solid", fgColor=GRAY)

        #################
        # Fill in table #
        #################

        # add title
        worksheet["A1"] = "Team Names"
        worksheet["A1"].font = headerColumnFont
        worksheet["A1"].fill = headerFill
        # add all team names
        for i, team in enumerate(self.teams):
            col = "A"
            worksheet[f"{col}{i + 2}"] = team.name
            worksheet[f"{col}{i + 2}"].font = teamNameFont
            worksheet[f"{col}{i + 2}"].fill = PatternFill(patternType="solid", fgColor=TEAM_ROW_COLORS[i])

        # add all stats
        statsWithTitles = self.statSheet(**kwargs).preferredOrderWithTitle()
        for row, teamId in enumerate([team.id for team in self.teams]):
            rowFill = PatternFill(patternType="solid", fgColor=TEAM_ROW_COLORS[row])
            for col, statWithTitle in enumerate(statsWithTitles):
                char = get_column_letter(col + 2)
                if row == 1:
                    # add stat header
                    worksheet[f"{char}{row}"] = statWithTitle[0]
                    worksheet[f"{char}{row}"].font = headerColumnFont
                    worksheet[f"{char}{row}"].fill = headerFill
                # add stat value
                worksheet[f"{char}{row + 2}"] = statWithTitle[1][teamId]
                worksheet[f"{char}{row + 2}"].fill = rowFill

        # put stats into table
        table = Table(displayName="YearStats",
                      ref="A1:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row))
        worksheet.add_table(table)

        # save
        workbook.save(filePath)
