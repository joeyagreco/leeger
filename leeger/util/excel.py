from __future__ import annotations

import os
import random
from datetime import datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from leeger.model.filter import AllTimeFilters, YearFilters
from leeger.model.league import Year, League
from leeger.util.excel_helper import (
    allTimeTeamsStatSheet,
    yearMatchupsStatSheet,
    allTimeMatchupsStatSheet,
)
from leeger.util.stat_sheet import yearStatSheet, leagueStatSheet


def leagueToExcel(league: League, filePath: Optional[str] = None, **kwargs) -> Workbook:
    """
    Saves the given League object to an Excel file.
    """
    overwrite = kwargs.pop("overwrite", False)

    if league is None:
        raise ValueError("'league' has not been set.")
    if filePath is not None:
        if os.path.exists(filePath) and not overwrite:
            raise FileExistsError(
                f"Cannot create file at path: '{filePath}' because there is already a file there."
            )
        else:
            try:
                os.remove(filePath)
            except FileNotFoundError:
                # we don't care if this doesn't exist
                pass

    # get owner names
    # and
    # make sure we have the same color for owners and their teams across sheets
    ownerIdToSeedMap = dict()
    ownerIds = list()
    ownerIdToNameMap = dict()
    for owner in league.owners:
        ownerIdToSeedMap[owner.id] = f"{owner.id}{datetime.now().date()}"
        ownerIds.append(owner.id)
        ownerIdToNameMap[owner.id] = owner.name
    ownerIdToColorMap = dict()
    for ownerId, seed in ownerIdToSeedMap.items():
        ownerIdToColorMap[ownerId] = _getRandomColor(0.5, seed)

    # gather info for all time teams to be used later
    allTimeTeamIds: list[str] = list()
    teamIdToColorMap = dict()
    workbook = None
    for year in league.years:
        for team in year.teams:
            allTimeTeamIds.append(team.id)
            teamIdToColorMap[team.id] = ownerIdToColorMap[team.ownerId]

        # add a sheet to the Excel document for this year
        workbook = _yearToExcel(year, workbook, **kwargs.copy())

    # add All-Time teams stats sheet
    # figure out index to put this sheet into
    # we want the sheets to be ordered: oldest year -> newest year -> all time teams -> all time matchups -> all time owners
    index = len(workbook.sheetnames)
    workbook.create_sheet("All Time Teams", index=index)
    worksheet = workbook["All Time Teams"]

    allTimeTeamsStatSheet_ = allTimeTeamsStatSheet(league, **kwargs.copy())

    allTimeFilters = AllTimeFilters.preferredOrderWithTitle(league, **kwargs.copy())
    _populateWorksheet(
        worksheet=worksheet,
        workbook=workbook,
        displayName="AllTimeTeamStats",
        titlesAndStatDicts=allTimeTeamsStatSheet_,
        entityIds=allTimeTeamIds,
        entityIdToColorMap=teamIdToColorMap,
        legendKeyValues=allTimeFilters,
        freezePanes="D2",
        saveToFilepath=filePath,
    )

    # add All-Time matchups stats sheet
    # figure out index to put this sheet into
    # we want the sheets to be ordered: oldest year -> newest year -> all time teams -> all time matchups -> all time owners
    index = len(workbook.sheetnames)
    workbook.create_sheet("All Time Matchups", index=index)
    worksheet = workbook["All Time Matchups"]

    (allTimeMatchupsStatSheet_, modifiedMatchupIdToOwnerIdMap) = allTimeMatchupsStatSheet(
        league, **kwargs.copy()
    )

    modifiedMatchupIdToColorMap: dict = dict()
    for modifiedMatchupId, ownerId in modifiedMatchupIdToOwnerIdMap.items():
        modifiedMatchupIdToColorMap[modifiedMatchupId] = ownerIdToColorMap[ownerId]

    allTimeFilters = AllTimeFilters.preferredOrderWithTitle(league, **kwargs.copy())
    _populateWorksheet(
        worksheet=worksheet,
        workbook=workbook,
        displayName="AllTimeMatchups",
        titlesAndStatDicts=allTimeMatchupsStatSheet_,
        entityIds=list(modifiedMatchupIdToOwnerIdMap.keys()),
        entityIdToColorMap=modifiedMatchupIdToColorMap,
        legendKeyValues=allTimeFilters,
        freezePanes="C2",
        saveToFilepath=filePath,
        boldColumnNumbers=[1, 2],
    )

    # add All-Time owner stats sheet
    # figure out index to put this sheet into
    # we want the sheets to be ordered: oldest year -> newest year -> all time teams -> all time matchups -> all time owners
    index = len(workbook.sheetnames)
    workbook.create_sheet("All Time Owners", index=index)
    worksheet = workbook["All Time Owners"]

    allTimeOwnerStatsWithTitles = leagueStatSheet(league, **kwargs.copy()).preferredOrderWithTitle()
    allTimeOwnerStatsWithTitles.insert(0, ("Owner", ownerIdToNameMap))

    return _populateWorksheet(
        worksheet=worksheet,
        workbook=workbook,
        displayName="AllTimeOwnerStats",
        titlesAndStatDicts=allTimeOwnerStatsWithTitles,
        entityIds=ownerIds,
        entityIdToColorMap=ownerIdToColorMap,
        legendKeyValues=allTimeFilters,
        freezePanes="B2",
        saveToFilepath=filePath,
    )


def _yearToExcel(year: Year, workbook: Optional[Workbook] = None, **kwargs) -> Workbook:
    """
    Saves the given Year object to an Excel File.
    If the given Excel file exists already, add a worksheet to it with this year.
    If the given Excel file does not exist, create a new workbook and add a worksheet to it with this year.
    """

    if workbook is None:
        # create workbook and sheet
        workbook = Workbook()
        workbook.create_sheet(f"{year.yearNumber} Teams", index=0)
        workbook.create_sheet(f"{year.yearNumber} Matchups", index=1)
        # remove default sheet
        del workbook["Sheet"]
    else:
        # figure out index to put this sheet into
        # we want the years as sheets in order from oldest -> newest year
        index = 0
        for i, sheetname in enumerate(workbook.sheetnames):
            if year.yearNumber > int(sheetname[:5]):
                index += 1
        workbook.create_sheet(f"{year.yearNumber} Teams", index=index)
        workbook.create_sheet(f"{year.yearNumber} Matchups", index=index + 1)
    worksheet = workbook[f"{year.yearNumber} Teams"]

    # get team names
    # and
    # make sure we have the same color for owners and their teams across sheets
    ownerIdToSeedMap = dict()
    teamIdToNameMap = dict()
    for team in year.teams:
        ownerIdToSeedMap[team.ownerId] = f"{team.ownerId}{datetime.now().date()}"
        teamIdToNameMap[team.id] = team.name
    ownerIdToColorMap = dict()
    for ownerId, seed in ownerIdToSeedMap.items():
        ownerIdToColorMap[ownerId] = _getRandomColor(0.5, seed)

    teamIdToColorMap = dict()
    teamIds = list()
    for team in year.teams:
        teamIdToColorMap[team.id] = ownerIdToColorMap[team.ownerId]
        teamIds.append(team.id)

    yearFilters = YearFilters.preferredOrderWithTitle(year, **kwargs.copy())

    yearStatsWithTitles = yearStatSheet(year, **kwargs.copy()).preferredOrderWithTitle()
    yearStatsWithTitles.insert(0, ("Team", teamIdToNameMap))
    # save Year teams to Excel sheet
    _populateWorksheet(
        worksheet=worksheet,
        workbook=workbook,
        displayName=f"Teams{year.yearNumber}",
        titlesAndStatDicts=yearStatsWithTitles,
        entityIds=teamIds,
        entityIdToColorMap=teamIdToColorMap,
        legendKeyValues=yearFilters,
        freezePanes="B2",
    )

    (yearMatchupsWithTitles, modifiedMatchupIdToOwnerIdMap) = yearMatchupsStatSheet(
        year, **kwargs.copy()
    )
    modifiedMatchupIdToColorMap: dict = dict()
    for modifiedMatchupId, ownerId in modifiedMatchupIdToOwnerIdMap.items():
        modifiedMatchupIdToColorMap[modifiedMatchupId] = ownerIdToColorMap[ownerId]
    worksheet = workbook[f"{year.yearNumber} Matchups"]
    # save Year matchups to Excel sheet
    _populateWorksheet(
        worksheet=worksheet,
        workbook=workbook,
        displayName=f"Matchups{year.yearNumber}",
        titlesAndStatDicts=yearMatchupsWithTitles,
        entityIds=list(modifiedMatchupIdToOwnerIdMap.keys()),
        entityIdToColorMap=modifiedMatchupIdToColorMap,
        legendKeyValues=yearFilters,
        freezePanes="B2",
    )
    return workbook


def _getRandomColor(tint: float = 0, seed: str = None) -> Color:
    """
    Used to get a random row color.
    """
    if seed:
        random.seed(seed)
    r = lambda: random.randint(0, 255)
    hexCode = "%02X%02X%02X" % (r(), r(), r())
    return Color(rgb=hexCode, tint=tint)


def _populateWorksheet(
    *,
    worksheet: Worksheet,
    workbook: Workbook,
    saveToFilepath: Optional[str] = None,
    displayName: str,
    titlesAndStatDicts: list[tuple[str, dict]],
    entityIds: list[str],
    entityIdToColorMap: dict[str, Color],
    legendKeyValues: list[tuple[str, Any]],
    freezePanes: str,
    boldColumnNumbers: Optional[list] = None,
) -> Workbook:
    """
    boldColumnNumbers is a list of column numbers whose values will be bolded.
        - Column numbers will be 1-indexed.
        - Default is [1]
    """
    boldColumnNumbers = [1] if boldColumnNumbers is None else boldColumnNumbers
    ####################
    # Styles for table #
    ####################

    # fonts
    HEADER_COLUMN_FONT = Font(size=12, bold=True)
    ENTITY_NAME_FONT = Font(size=11, bold=True)

    # colors
    BLACK = Color(rgb="000000")
    LIGHT_GRAY = Color(rgb="B8B8B8")
    MEDIUM_GRAY = Color(rgb="828282")

    # fills
    HEADER_FILL = PatternFill(patternType="solid", fgColor=LIGHT_GRAY)

    #################
    # Fill in table #
    #################

    # add all stats
    for rowNumber, entityId in enumerate(entityIds):
        rowFill = PatternFill(patternType="solid", fgColor=entityIdToColorMap[entityId])
        for columnNumber, (title, statDict) in enumerate(titlesAndStatDicts):
            char = get_column_letter(columnNumber + 1)
            if rowNumber == 1:
                # add stat header
                cell = f"{char}{rowNumber}"
                worksheet[cell] = title
                worksheet[cell].font = HEADER_COLUMN_FONT
                worksheet[cell].fill = HEADER_FILL
                worksheet[cell].alignment = Alignment(horizontal="center")
            # add stat value
            cell = f"{char}{rowNumber + 2}"
            if entityId in statDict:
                worksheet[cell] = statDict[entityId]
            else:
                worksheet[cell] = "N/A"
            worksheet[cell].fill = rowFill
            if columnNumber + 1 in boldColumnNumbers:
                worksheet[cell].font = ENTITY_NAME_FONT

    # put stats into table
    table = Table(
        displayName=displayName,
        ref="A1:" + get_column_letter(worksheet.max_column) + str(len(entityIds) + 1),
    )
    worksheet.add_table(table)
    # freeze owner name column and header row
    worksheet.freeze_panes = freezePanes

    # add legend for filters
    # define border to go around legend
    topSideSolid = Side(border_style="thick", color=BLACK)
    bottomSideSolid = Side(border_style="thick", color=BLACK)
    bottomSideThin = Side(border_style="thin", color=BLACK)
    leftSideSolid = Side(border_style="thick", color=BLACK)
    rightSideSolid = Side(border_style="thick", color=BLACK)
    # legend formatting
    legendCellAlignment = Alignment(horizontal="center")

    legendRowNumber = len(entityIds) + 4
    legendColLetter = "A"
    titleCell = f"A{legendRowNumber}"
    worksheet[titleCell] = "Filters Applied"
    worksheet[titleCell].fill = PatternFill(patternType="solid", fgColor=MEDIUM_GRAY)
    worksheet[titleCell].font = Font(bold=True)
    border = Border(
        left=leftSideSolid, right=rightSideSolid, top=topSideSolid, bottom=bottomSideThin
    )
    worksheet[titleCell].border = border
    worksheet[titleCell].alignment = legendCellAlignment

    for kwarg_title, kwarg_value in legendKeyValues:
        legendRowNumber += 1
        cell = f"{legendColLetter}{legendRowNumber}"
        worksheet[cell] = f"{kwarg_title}: {kwarg_value}"
        worksheet[cell].fill = PatternFill(patternType="solid", fgColor=MEDIUM_GRAY)
        worksheet[cell].alignment = legendCellAlignment
        border = Border(left=leftSideSolid, right=rightSideSolid)
        worksheet[cell].border = border

    # add border to last cell
    border = Border(left=leftSideSolid, right=rightSideSolid, bottom=bottomSideSolid)
    worksheet[f"{legendColLetter}{legendRowNumber}"].border = border

    # set column widths
    TITLE_MULTIPLIER = 1.2
    NAME_MULTIPLIER = 1.0
    DATA_MULTIPLIER = 0.8
    dim_holder = DimensionHolder(worksheet=worksheet)

    for columnNumber in range(worksheet.min_column, worksheet.max_column + 1):
        # figure out the width we want this column to be
        # first column has entity names, so use a different multiplier for them
        isNameColumn = columnNumber == 1
        maxWidth = 0
        for rowNumber, cell in enumerate(worksheet[get_column_letter(columnNumber)]):
            if cell.value:
                # count title/name cell characters as more than a data cell
                if rowNumber == 0:
                    multiplier = TITLE_MULTIPLIER
                elif isNameColumn:
                    multiplier = NAME_MULTIPLIER
                else:
                    multiplier = DATA_MULTIPLIER
                maxWidth = max((maxWidth, len(str(cell.value)) * multiplier))
        dim_holder[get_column_letter(columnNumber)] = ColumnDimension(
            worksheet, min=columnNumber, max=columnNumber, width=maxWidth + 7
        )

    worksheet.column_dimensions = dim_holder

    # save
    if saveToFilepath is not None:
        workbook.save(saveToFilepath)
    return workbook
