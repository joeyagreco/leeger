import os
import random
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from leeger.model.abstract.UniqueId import UniqueId
from leeger.model.league.Team import Team
from leeger.model.league.Week import Week
from leeger.model.stat.YearStatSheet import YearStatSheet


@dataclass(kw_only=True)
class Year(UniqueId):
    yearNumber: int
    teams: list[Team]
    weeks: list[Week]

    def statSheet(self, **kwargs) -> YearStatSheet:
        from leeger.calculator.year_calculator.AWALYearCalculator import AWALYearCalculator
        from leeger.calculator.year_calculator.GameOutcomeYearCalculator import GameOutcomeYearCalculator
        from leeger.calculator.year_calculator.PlusMinusYearCalculator import PlusMinusYearCalculator
        from leeger.calculator.year_calculator.PointsScoredYearCalculator import PointsScoredYearCalculator
        from leeger.calculator.year_calculator.ScoringShareYearCalculator import ScoringShareYearCalculator
        from leeger.calculator.year_calculator.ScoringStandardDeviationYearCalculator import \
            ScoringStandardDeviationYearCalculator
        from leeger.calculator.year_calculator.SingleScoreYearCalculator import SingleScoreYearCalculator
        from leeger.calculator.year_calculator.SmartWinsYearCalculator import SmartWinsYearCalculator
        from leeger.calculator.year_calculator.SSLYearCalculator import SSLYearCalculator

        # Game Outcome
        wins = GameOutcomeYearCalculator.getWins(self, **kwargs)
        losses = GameOutcomeYearCalculator.getLosses(self, **kwargs)
        ties = GameOutcomeYearCalculator.getTies(self, **kwargs)
        winPercentage = GameOutcomeYearCalculator.getWinPercentage(self, **kwargs)
        wal = GameOutcomeYearCalculator.getWAL(self, **kwargs)
        walPerGame = GameOutcomeYearCalculator.getWALPerGame(self, **kwargs)

        # AWAL
        awal = AWALYearCalculator.getAWAL(self, **kwargs)
        awalPerGame = AWALYearCalculator.getAWALPerGame(self, **kwargs)
        opponentAWAL = AWALYearCalculator.getOpponentAWAL(self, **kwargs)
        opponentAWALPerGame = AWALYearCalculator.getOpponentAWALPerGame(self, **kwargs)

        # Smart Wins
        smartWins = SmartWinsYearCalculator.getSmartWins(self, **kwargs)
        smartWinsPerGame = SmartWinsYearCalculator.getSmartWinsPerGame(self, **kwargs)
        opponentSmartWins = SmartWinsYearCalculator.getOpponentSmartWins(self, **kwargs)
        opponentSmartWinsPerGame = SmartWinsYearCalculator.getOpponentSmartWinsPerGame(self, **kwargs)

        # Points Scored
        pointsScored = PointsScoredYearCalculator.getPointsScored(self, **kwargs)
        pointsScoredPerGame = PointsScoredYearCalculator.getPointsScoredPerGame(self, **kwargs)
        opponentPointsScored = PointsScoredYearCalculator.getOpponentPointsScored(self, **kwargs)
        opponentPointsScoredPerGame = PointsScoredYearCalculator.getOpponentPointsScoredPerGame(self, **kwargs)

        # Scoring Share
        scoringShare = ScoringShareYearCalculator.getScoringShare(self, **kwargs)
        opponentScoringShare = ScoringShareYearCalculator.getOpponentScoringShare(self, **kwargs)

        # Single Score
        maxScore = SingleScoreYearCalculator.getMaxScore(self, **kwargs)
        minScore = SingleScoreYearCalculator.getMinScore(self, **kwargs)

        # Scoring Standard Deviation
        scoringStandardDeviation = ScoringStandardDeviationYearCalculator.getScoringStandardDeviation(self, **kwargs)

        # Plus Minus
        plusMinus = PlusMinusYearCalculator.getPlusMinus(self, **kwargs)

        # SSL
        teamScore = SSLYearCalculator.getTeamScore(self, **kwargs)
        teamSuccess = SSLYearCalculator.getTeamSuccess(self, **kwargs)
        teamLuck = SSLYearCalculator.getTeamLuck(self, **kwargs)

        return YearStatSheet(wins=wins, losses=losses, ties=ties, winPercentage=winPercentage, wal=wal,
                             walPerGame=walPerGame, awal=awal, awalPerGame=awalPerGame, opponentAWAL=opponentAWAL,
                             opponentAWALPerGame=opponentAWALPerGame, smartWins=smartWins,
                             smartWinsPerGame=smartWinsPerGame, opponentSmartWins=opponentSmartWins,
                             opponentSmartWinsPerGame=opponentSmartWinsPerGame, pointsScored=pointsScored,
                             pointsScoredPerGame=pointsScoredPerGame, opponentPointsScored=opponentPointsScored,
                             opponentPointsScoredPerGame=opponentPointsScoredPerGame, scoringShare=scoringShare,
                             opponentScoringShare=opponentScoringShare, maxScore=maxScore, minScore=minScore,
                             scoringStandardDeviation=scoringStandardDeviation, plusMinus=plusMinus,
                             teamScore=teamScore, teamSuccess=teamSuccess, teamLuck=teamLuck)

    def toExcel(self, filePath: str, **kwargs) -> None:
        """
        If the given Excel file exists already, add a worksheet to it with this year.
        If the given Excel file does not exist, create a new workbook and add a worksheet to it with this year.
        """

        if os.path.exists(filePath):
            workbook = load_workbook(filename=filePath)
            # figure out index to put this sheet into
            # we want the years as sheets in order from oldest -> newest year
            index = 0
            for i, sheetname in enumerate(workbook.sheetnames):
                if self.yearNumber > int(sheetname):
                    index += 1
            workbook.create_sheet(str(self.yearNumber), index=index)
        else:
            # create workbook and sheet
            workbook = Workbook()
            workbook.create_sheet(str(self.yearNumber), index=0)
            # remove default sheet
            del workbook["Sheet"]
        worksheet = workbook[str(self.yearNumber)]

        ####################
        # Styles for table #
        ####################

        # fonts
        headerColumnFont = Font(size=12, bold=True)
        teamNameFont = Font(size=11, bold=True)

        # colors
        GRAY = Color(rgb="B8B8B8")

        def getRandomColor(tint: float = 0) -> Color:
            r = lambda: random.randint(0, 255)
            hexCode = "%02X%02X%02X" % (r(), r(), r())
            return Color(rgb=hexCode, tint=tint)

        TEAM_ROW_COLORS = [getRandomColor(0.5) for _ in range(len(self.teams))]

        # fills
        headerFill = PatternFill(patternType="solid", fgColor=GRAY)

        #################
        # Fill in table #
        #################

        # add title
        worksheet["A1"] = "Team Names"
        worksheet["A1"].font = headerColumnFont
        worksheet["A1"].fill = headerFill
        # add all team names
        for i, team in enumerate(self.teams):
            col = "A"
            worksheet[f"{col}{i + 2}"] = team.name
            worksheet[f"{col}{i + 2}"].font = teamNameFont
            worksheet[f"{col}{i + 2}"].fill = PatternFill(patternType="solid", fgColor=TEAM_ROW_COLORS[i])

        # add all stats
        statsWithTitles = self.statSheet(**kwargs).preferredOrderWithTitle()
        for row, teamId in enumerate([team.id for team in self.teams]):
            rowFill = PatternFill(patternType="solid", fgColor=TEAM_ROW_COLORS[row])
            for col, statWithTitle in enumerate(statsWithTitles):
                char = get_column_letter(col + 2)
                if row == 1:
                    # add stat header
                    worksheet[f"{char}{row}"] = statWithTitle[0]
                    worksheet[f"{char}{row}"].font = headerColumnFont
                    worksheet[f"{char}{row}"].fill = headerFill
                # add stat value
                worksheet[f"{char}{row + 2}"] = statWithTitle[1][teamId]
                worksheet[f"{char}{row + 2}"].fill = rowFill

        # put stats into table
        table = Table(displayName=f"YearStats{self.yearNumber}",
                      ref="A1:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row))
        worksheet.add_table(table)

        # save
        workbook.save(filePath)
