from __future__ import annotations

import os
import random
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from leeger.model.abstract.UniqueId import UniqueId
from leeger.model.league.Owner import Owner
from leeger.model.league.Year import Year
from leeger.model.stat.AllTimeStatSheet import AllTimeStatSheet


@dataclass(kw_only=True)
class League(UniqueId):
    name: str
    owners: list[Owner]
    years: list[Year]

    def __add__(self, otherLeague: League) -> League:
        """
        Combines *this* League with the given League.
        The combined League will be validated before it is returned.

        Special behaviors:
            - name
                - "name" will become a combination of both league's names IF the names are not the same.
            - owners
                - "owners" will be merged on Owner.name, since this field must be unique by League.
                - Unmerged owners will simply be combined.
            - years
                - "years" will be combined in order oldestYearNumber -> newestYearNumber.
                - Duplicate Year.yearNumber across leagues will raise an exception.
        """
        from leeger.decorator.validate.common import leagueValidation
        # first, validate the leagues we want to combine.
        leagueValidation.runAllChecks(self)
        leagueValidation.runAllChecks(otherLeague)

        newName = f"'{self.name}' + '{otherLeague.name}' League" if self.name != otherLeague.name else self.name
        newOwners = list()
        newYears = list()

        # merge/combine owners
        thisLeagueOwnerNames = [owner.name for owner in self.owners]
        otherLeagueOwnerNames = [owner.name for owner in otherLeague.owners]

        for ownerName in thisLeagueOwnerNames:
            if ownerName in otherLeagueOwnerNames:
                otherLeagueOwnerNames.remove(ownerName)
            newOwners.append(Owner(name=ownerName))
        for ownerName in otherLeagueOwnerNames:
            newOwners.append(Owner(name=ownerName))

        # combine years in order
        if self.years[-1].yearNumber < otherLeague.years[0].yearNumber:
            # order of years goes *this* League -> otherLeague
            newYears = self.years + otherLeague.years
        else:
            # order of years goes otherLeague -> *this* League
            newYears = otherLeague.years + self.years

        newLeague = League(name=newName, owners=newOwners, years=newYears)

        # validate new league
        leagueValidation.runAllChecks(newLeague)
        return newLeague

    def statSheet(self, **kwargs) -> AllTimeStatSheet:
        from leeger.calculator.all_time_calculator.AWALAllTimeCalculator import AWALAllTimeCalculator
        from leeger.calculator.all_time_calculator.GameOutcomeAllTimeCalculator import GameOutcomeAllTimeCalculator
        from leeger.calculator.all_time_calculator.PlusMinusAllTimeCalculator import PlusMinusAllTimeCalculator
        from leeger.calculator.all_time_calculator.PointsScoredAllTimeCalculator import \
            PointsScoredAllTimeCalculator
        from leeger.calculator.all_time_calculator.SSLAllTimeCalculator import SSLAllTimeCalculator
        from leeger.calculator.all_time_calculator.ScoringShareAllTimeCalculator import \
            ScoringShareAllTimeCalculator
        from leeger.calculator.all_time_calculator.ScoringStandardDeviationAllTimeCalculator import \
            ScoringStandardDeviationAllTimeCalculator
        from leeger.calculator.all_time_calculator.SingleScoreAllTimeCalculator import SingleScoreAllTimeCalculator
        from leeger.calculator.all_time_calculator.SmartWinsAllTimeCalculator import SmartWinsAllTimeCalculator

        # Game Outcome
        wins = GameOutcomeAllTimeCalculator.getWins(self, **kwargs)
        losses = GameOutcomeAllTimeCalculator.getLosses(self, **kwargs)
        ties = GameOutcomeAllTimeCalculator.getTies(self, **kwargs)
        winPercentage = GameOutcomeAllTimeCalculator.getWinPercentage(self, **kwargs)
        wal = GameOutcomeAllTimeCalculator.getWAL(self, **kwargs)
        walPerGame = GameOutcomeAllTimeCalculator.getWALPerGame(self, **kwargs)

        # AWAL
        awal = AWALAllTimeCalculator.getAWAL(self, **kwargs)
        awalPerGame = AWALAllTimeCalculator.getAWALPerGame(self, **kwargs)
        opponentAWAL = AWALAllTimeCalculator.getOpponentAWAL(self, **kwargs)
        opponentAWALPerGame = AWALAllTimeCalculator.getOpponentAWALPerGame(self, **kwargs)

        # Smart Wins
        smartWins = SmartWinsAllTimeCalculator.getSmartWins(self, **kwargs)
        smartWinsPerGame = SmartWinsAllTimeCalculator.getSmartWinsPerGame(self, **kwargs)
        opponentSmartWins = SmartWinsAllTimeCalculator.getOpponentSmartWins(self, **kwargs)
        opponentSmartWinsPerGame = SmartWinsAllTimeCalculator.getOpponentSmartWinsPerGame(self, **kwargs)

        # Points Scored
        pointsScored = PointsScoredAllTimeCalculator.getPointsScored(self, **kwargs)
        pointsScoredPerGame = PointsScoredAllTimeCalculator.getPointsScoredPerGame(self, **kwargs)
        opponentPointsScored = PointsScoredAllTimeCalculator.getOpponentPointsScored(self, **kwargs)
        opponentPointsScoredPerGame = PointsScoredAllTimeCalculator.getOpponentPointsScoredPerGame(self, **kwargs)

        # Scoring Share
        scoringShare = ScoringShareAllTimeCalculator.getScoringShare(self, **kwargs)
        opponentScoringShare = ScoringShareAllTimeCalculator.getOpponentScoringShare(self, **kwargs)

        # Single Score
        maxScore = SingleScoreAllTimeCalculator.getMaxScore(self, **kwargs)
        minScore = SingleScoreAllTimeCalculator.getMinScore(self, **kwargs)

        # Scoring Standard Deviation
        scoringStandardDeviation = ScoringStandardDeviationAllTimeCalculator.getScoringStandardDeviation(self, **kwargs)

        # Plus Minus
        plusMinus = PlusMinusAllTimeCalculator.getPlusMinus(self, **kwargs)

        # SSL
        teamScore = SSLAllTimeCalculator.getTeamScore(self, **kwargs)
        teamSuccess = SSLAllTimeCalculator.getTeamSuccess(self, **kwargs)
        teamLuck = SSLAllTimeCalculator.getTeamLuck(self, **kwargs)

        return AllTimeStatSheet(wins=wins, losses=losses, ties=ties, winPercentage=winPercentage, wal=wal,
                                walPerGame=walPerGame, awal=awal, awalPerGame=awalPerGame, opponentAWAL=opponentAWAL,
                                opponentAWALPerGame=opponentAWALPerGame, smartWins=smartWins,
                                smartWinsPerGame=smartWinsPerGame, opponentSmartWins=opponentSmartWins,
                                opponentSmartWinsPerGame=opponentSmartWinsPerGame, pointsScored=pointsScored,
                                pointsScoredPerGame=pointsScoredPerGame, opponentPointsScored=opponentPointsScored,
                                opponentPointsScoredPerGame=opponentPointsScoredPerGame, scoringShare=scoringShare,
                                opponentScoringShare=opponentScoringShare, maxScore=maxScore, minScore=minScore,
                                scoringStandardDeviation=scoringStandardDeviation, plusMinus=plusMinus,
                                teamScore=teamScore, teamSuccess=teamSuccess, teamLuck=teamLuck)

    def toExcel(self, filePath: str, **kwargs) -> None:
        if os.path.exists(filePath):
            raise FileExistsError(f"Cannot create file at path: '{filePath}' because there is already a file there.")
        for year in self.years:
            year.toExcel(filePath, **kwargs)

        # add All-Time stats sheet
        workbook = load_workbook(filename=filePath)
        # figure out index to put this sheet into
        # we want the sheets to be ordered: oldest year -> newest year -> all time
        index = len(workbook.sheetnames)
        workbook.create_sheet("All Time", index=index)
        worksheet = workbook["All Time"]

        ####################
        # Styles for table #
        ####################

        # fonts
        headerColumnFont = Font(size=12, bold=True)
        teamNameFont = Font(size=11, bold=True)

        # colors
        GRAY = Color(rgb="B8B8B8")

        def getRandomColor(tint: float = 0) -> Color:
            r = lambda: random.randint(0, 255)
            hexCode = "%02X%02X%02X" % (r(), r(), r())
            return Color(rgb=hexCode, tint=tint)

        OWNER_ROW_COLORS = [getRandomColor(0.5) for _ in range(len(self.owners))]

        # fills
        headerFill = PatternFill(patternType="solid", fgColor=GRAY)

        #################
        # Fill in table #
        #################

        # add title
        worksheet["A1"] = "Owner Names"
        worksheet["A1"].font = headerColumnFont
        worksheet["A1"].fill = headerFill
        # add all team names
        for i, owner in enumerate(self.owners):
            col = "A"
            worksheet[f"{col}{i + 2}"] = owner.name
            worksheet[f"{col}{i + 2}"].font = teamNameFont
            worksheet[f"{col}{i + 2}"].fill = PatternFill(patternType="solid", fgColor=OWNER_ROW_COLORS[i])

        # add all stats
        statsWithTitles = self.statSheet(**kwargs).preferredOrderWithTitle()
        for row, ownerId in enumerate([owner.id for owner in self.owners]):
            rowFill = PatternFill(patternType="solid", fgColor=OWNER_ROW_COLORS[row])
            for col, statWithTitle in enumerate(statsWithTitles):
                char = get_column_letter(col + 2)
                if row == 1:
                    # add stat header
                    worksheet[f"{char}{row}"] = statWithTitle[0]
                    worksheet[f"{char}{row}"].font = headerColumnFont
                    worksheet[f"{char}{row}"].fill = headerFill
                # add stat value
                worksheet[f"{char}{row + 2}"] = statWithTitle[1][ownerId]
                worksheet[f"{char}{row + 2}"].fill = rowFill

        # put stats into table
        table = Table(displayName=f"AllTimeStats",
                      ref="A1:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row))
        worksheet.add_table(table)

        # save
        workbook.save(filePath)
