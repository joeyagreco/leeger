import os
import tempfile
import unittest
from decimal import Decimal

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from leeger.model.league.League import League
from leeger.model.league.Matchup import Matchup
from leeger.model.league.Team import Team
from leeger.model.league.Week import Week
from leeger.model.league.Year import Year
from leeger.util.excel import leagueToExcel
from test.helper.prototypes import getNDefaultOwnersAndTeams


class TestExcel(unittest.TestCase):
    def test_leagueToExcel_filePathGiven_happyPath(self):
        owners, teams1 = getNDefaultOwnersAndTeams(2)
        teams1[0].name = "a"
        teams1[1].name = "b"
        matchup1 = Matchup(teamAId=teams1[0].id, teamBId=teams1[1].id, teamAScore=1, teamBScore=2)
        week1 = Week(weekNumber=1, matchups=[matchup1])
        year1 = Year(yearNumber=2000, teams=teams1, weeks=[week1])

        teams2 = [Team(ownerId=owners[0].id, name="a2"), Team(ownerId=owners[1].id, name="b2")]
        matchup2 = Matchup(teamAId=teams2[0].id, teamBId=teams2[1].id, teamAScore=1, teamBScore=2)
        week2 = Week(weekNumber=1, matchups=[matchup2])
        year2 = Year(yearNumber=2001, teams=teams2, weeks=[week2])

        teams3 = [Team(ownerId=owners[0].id, name="a3"), Team(ownerId=owners[1].id, name="b3")]
        matchup3 = Matchup(teamAId=teams3[0].id, teamBId=teams3[1].id, teamAScore=1, teamBScore=2)
        week3 = Week(weekNumber=1, matchups=[matchup3])
        year3 = Year(yearNumber=2002, teams=teams3, weeks=[week3])

        league = League(name="", owners=owners, years=[year1, year2, year3])

        with tempfile.TemporaryDirectory() as tempDir:
            fullPath = os.path.join(tempDir, "tmp.xlsx")
            leagueToExcel(league, fullPath)

            # open created Excel file to check that it was saved correctly
            workbook = load_workbook(filename=fullPath)

            # check sheets
            self.assertEqual(9, len(workbook.sheetnames))
            self.assertEqual("2000 Teams", workbook.sheetnames[0])
            self.assertEqual("2000 Matchups", workbook.sheetnames[1])
            self.assertEqual("2001 Teams", workbook.sheetnames[2])
            self.assertEqual("2001 Matchups", workbook.sheetnames[3])
            self.assertEqual("2002 Teams", workbook.sheetnames[4])
            self.assertEqual("2002 Matchups", workbook.sheetnames[5])
            self.assertEqual("All Time Teams", workbook.sheetnames[6])
            self.assertEqual("All Time Matchups", workbook.sheetnames[7])
            self.assertEqual("All Time Owners", workbook.sheetnames[8])

            # check sheet values
            # test year worksheet values
            worksheet2000TeamValues = [
                [
                    "Team",
                    "Games Played",
                    "Wins",
                    "Losses",
                    "Ties",
                    "Win Percentage",
                    "WAL",
                    "WAL Per Game",
                    "AWAL",
                    "AWAL Per Game",
                    "Opponent AWAL",
                    "Opponent AWAL Per Game",
                    "Smart Wins",
                    "Smart Wins Per Game",
                    "Opponent Smart Wins",
                    "Opponent Smart Wins Per Game",
                    "Points Scored",
                    "Points Scored Per Game",
                    "Opponent Points Scored",
                    "Opponent Points Scored Per Game",
                    "Scoring Share",
                    "Opponent Scoring Share",
                    "Max Scoring Share",
                    "Min Scoring Share",
                    "Max Score",
                    "Min Score",
                    "Scoring Standard Deviation",
                    "Plus/Minus",
                    "Team Score",
                    "Team Success",
                    "Team Luck",
                ],
                [
                    "a",
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b",
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
            ]

            worksheet2000MatchupValues = [
                [
                    "Team For",
                    "Team Against",
                    "Week Number",
                    "Matchup Type",
                    "Points For",
                    "Points Against",
                ],
                ["a", "b", 1, "REGULAR_SEASON", 1, 2],
                ["b", "a", 1, "REGULAR_SEASON", 2, 1],
            ]

            worksheet2001TeamValues = [
                [
                    "Team",
                    "Games Played",
                    "Wins",
                    "Losses",
                    "Ties",
                    "Win Percentage",
                    "WAL",
                    "WAL Per Game",
                    "AWAL",
                    "AWAL Per Game",
                    "Opponent AWAL",
                    "Opponent AWAL Per Game",
                    "Smart Wins",
                    "Smart Wins Per Game",
                    "Opponent Smart Wins",
                    "Opponent Smart Wins Per Game",
                    "Points Scored",
                    "Points Scored Per Game",
                    "Opponent Points Scored",
                    "Opponent Points Scored Per Game",
                    "Scoring Share",
                    "Opponent Scoring Share",
                    "Max Scoring Share",
                    "Min Scoring Share",
                    "Max Score",
                    "Min Score",
                    "Scoring Standard Deviation",
                    "Plus/Minus",
                    "Team Score",
                    "Team Success",
                    "Team Luck",
                ],
                [
                    "a2",
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b2",
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
            ]

            worksheet2001MatchupValues = [
                [
                    "Team For",
                    "Team Against",
                    "Week Number",
                    "Matchup Type",
                    "Points For",
                    "Points Against",
                ],
                ["a2", "b2", 1, "REGULAR_SEASON", 1, 2],
                ["b2", "a2", 1, "REGULAR_SEASON", 2, 1],
            ]

            worksheet2002TeamValues = [
                [
                    "Team",
                    "Games Played",
                    "Wins",
                    "Losses",
                    "Ties",
                    "Win Percentage",
                    "WAL",
                    "WAL Per Game",
                    "AWAL",
                    "AWAL Per Game",
                    "Opponent AWAL",
                    "Opponent AWAL Per Game",
                    "Smart Wins",
                    "Smart Wins Per Game",
                    "Opponent Smart Wins",
                    "Opponent Smart Wins Per Game",
                    "Points Scored",
                    "Points Scored Per Game",
                    "Opponent Points Scored",
                    "Opponent Points Scored Per Game",
                    "Scoring Share",
                    "Opponent Scoring Share",
                    "Max Scoring Share",
                    "Min Scoring Share",
                    "Max Score",
                    "Min Score",
                    "Scoring Standard Deviation",
                    "Plus/Minus",
                    "Team Score",
                    "Team Success",
                    "Team Luck",
                ],
                [
                    "a3",
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b3",
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
            ]

            worksheet2002MatchupValues = [
                [
                    "Team For",
                    "Team Against",
                    "Week Number",
                    "Matchup Type",
                    "Points For",
                    "Points Against",
                ],
                ["a3", "b3", 1, "REGULAR_SEASON", 1, 2],
                ["b3", "a3", 1, "REGULAR_SEASON", 2, 1],
            ]

            yearWorksheetsAndValues = [
                (workbook["2000 Teams"], worksheet2000TeamValues),
                (workbook["2000 Matchups"], worksheet2000MatchupValues),
                (workbook["2001 Teams"], worksheet2001TeamValues),
                (workbook["2001 Matchups"], worksheet2001MatchupValues),
                (workbook["2002 Teams"], worksheet2002TeamValues),
                (workbook["2002 Matchups"], worksheet2002MatchupValues),
            ]

            for i, (currentWorksheet, allWorksheetValues) in enumerate(yearWorksheetsAndValues):
                for j, worksheetValues in enumerate(allWorksheetValues):
                    for k, worksheetValue in enumerate(worksheetValues):
                        row = j + 1
                        column = get_column_letter(k + 1)
                        self.assertEqual(worksheetValue, currentWorksheet[f"{column}{row}"].value)
                # check that "next" cell is empty
                self.assertIsNone(currentWorksheet["A4"].value)
                # check legend
                self.assertEqual("Filters Applied", currentWorksheet["A6"].value)
                self.assertEqual("Week Number Start: 1", currentWorksheet["A7"].value)
                self.assertEqual("Week Number End: 1", currentWorksheet["A8"].value)
                self.assertEqual("Only Regular Season: False", currentWorksheet["A9"].value)
                self.assertEqual("Only Post Season: False", currentWorksheet["A10"].value)
                self.assertEqual("Only Championship: False", currentWorksheet["A11"].value)
                self.assertEqual("Include Multi-Week Matchups: True", currentWorksheet["A12"].value)
                self.assertIsNone(currentWorksheet["A13"].value)

            # test all time teams worksheet values
            allTimeTeamsWorksheet = workbook["All Time Teams"]

            allTimeTeamsValues = [
                [
                    "Team",
                    "Owner",
                    "Year",
                    "Games Played",
                    "Wins",
                    "Losses",
                    "Ties",
                    "Win Percentage",
                    "WAL",
                    "WAL Per Game",
                    "AWAL",
                    "AWAL Per Game",
                    "Opponent AWAL",
                    "Opponent AWAL Per Game",
                    "Smart Wins",
                    "Smart Wins Per Game",
                    "Opponent Smart Wins",
                    "Opponent Smart Wins Per Game",
                    "Points Scored",
                    "Points Scored Per Game",
                    "Opponent Points Scored",
                    "Opponent Points Scored Per Game",
                    "Scoring Share",
                    "Opponent Scoring Share",
                    "Max Scoring Share",
                    "Min Scoring Share",
                    "Max Score",
                    "Min Score",
                    "Scoring Standard Deviation",
                    "Plus/Minus",
                    "Team Score",
                    "Team Success",
                    "Team Luck",
                ],
                [
                    "a",
                    "1",
                    2000,
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b",
                    "2",
                    2000,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
                [
                    "a2",
                    "1",
                    2001,
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b2",
                    "2",
                    2001,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
                [
                    "a3",
                    "1",
                    2002,
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b3",
                    "2",
                    2002,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
            ]

            for rowNumber in range(1, 8):
                values = allTimeTeamsValues[rowNumber - 1]
                for columnNumber, value in enumerate(values):
                    cell = f"{get_column_letter(columnNumber + 1)}{rowNumber}"
                    self.assertEqual(values[columnNumber], allTimeTeamsWorksheet[cell].value)
            # check that "next" cell is empty
            self.assertIsNone(allTimeTeamsWorksheet["A8"].value)
            # check legend
            self.assertEqual("Filters Applied", allTimeTeamsWorksheet["A10"].value)
            self.assertEqual("Week Number Start: 1", allTimeTeamsWorksheet["A11"].value)
            self.assertEqual("Year Number Start: 2000", allTimeTeamsWorksheet["A12"].value)
            self.assertEqual("Week Number End: 1", allTimeTeamsWorksheet["A13"].value)
            self.assertEqual("Year Number End: 2002", allTimeTeamsWorksheet["A14"].value)
            self.assertEqual("Only Regular Season: False", allTimeTeamsWorksheet["A15"].value)
            self.assertEqual("Only Post Season: False", allTimeTeamsWorksheet["A16"].value)
            self.assertEqual("Only Championship: False", allTimeTeamsWorksheet["A17"].value)
            self.assertIsNone(allTimeTeamsWorksheet["A18"].value)

            # test all time matchups worksheet values
            allTimeMatchupsWorksheet = workbook["All Time Matchups"]

            allTimeMatchupsValues = [
                [
                    "Team For",
                    "Owner For",
                    "Team Against",
                    "Owner Against",
                    "Year",
                    "Week Number",
                    "Matchup Type",
                    "Points For",
                    "Points Against",
                ],
                ["a", "1", "b", "2", 2000, 1, "REGULAR_SEASON", 1, 2],
                ["b", "2", "a", "1", 2000, 1, "REGULAR_SEASON", 2, 1],
                ["a2", "1", "b2", "2", 2001, 1, "REGULAR_SEASON", 1, 2],
                ["b2", "2", "a2", "1", 2001, 1, "REGULAR_SEASON", 2, 1],
                ["a3", "1", "b3", "2", 2002, 1, "REGULAR_SEASON", 1, 2],
                ["b3", "2", "a3", "1", 2002, 1, "REGULAR_SEASON", 2, 1],
            ]

            for rowNumber in range(1, 8):
                values = allTimeMatchupsValues[rowNumber - 1]
                for columnNumber, value in enumerate(values):
                    cell = f"{get_column_letter(columnNumber + 1)}{rowNumber}"
                    self.assertEqual(values[columnNumber], allTimeMatchupsWorksheet[cell].value)
            # check that "next" cell is empty
            self.assertIsNone(allTimeMatchupsWorksheet["A8"].value)
            # check legend
            self.assertEqual("Filters Applied", allTimeMatchupsWorksheet["A10"].value)
            self.assertEqual("Week Number Start: 1", allTimeMatchupsWorksheet["A11"].value)
            self.assertEqual("Year Number Start: 2000", allTimeMatchupsWorksheet["A12"].value)
            self.assertEqual("Week Number End: 1", allTimeMatchupsWorksheet["A13"].value)
            self.assertEqual("Year Number End: 2002", allTimeMatchupsWorksheet["A14"].value)
            self.assertEqual("Only Regular Season: False", allTimeMatchupsWorksheet["A15"].value)
            self.assertEqual("Only Post Season: False", allTimeMatchupsWorksheet["A16"].value)
            self.assertEqual("Only Championship: False", allTimeMatchupsWorksheet["A17"].value)
            self.assertIsNone(allTimeMatchupsWorksheet["A18"].value)

            # test all time owners worksheet values
            allTimeOwnersWorksheet = workbook["All Time Owners"]

            allTimeOwnersValues = [
                [
                    "Owner",
                    "Games Played",
                    "Wins",
                    "Losses",
                    "Ties",
                    "Win Percentage",
                    "WAL",
                    "WAL Per Game",
                    "AWAL",
                    "AWAL Per Game",
                    "Opponent AWAL",
                    "Opponent AWAL Per Game",
                    "Smart Wins",
                    "Smart Wins Per Game",
                    "Opponent Smart Wins",
                    "Opponent Smart Wins Per Game",
                    "Points Scored",
                    "Points Scored Per Game",
                    "Opponent Points Scored",
                    "Opponent Points Scored Per Game",
                    "Scoring Share",
                    "Opponent Scoring Share",
                    "Max Scoring Share",
                    "Min Scoring Share",
                    "Max Score",
                    "Min Score",
                    "Scoring Standard Deviation",
                    "Plus/Minus",
                    "Adjusted Team Score",
                    "Adjusted Team Success",
                    "Adjusted Team Luck",
                ],
                [
                    "1",
                    3,
                    0,
                    3,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    3,
                    1,
                    0.6,
                    0.2,
                    2.4,
                    0.8,
                    3,
                    1,
                    6,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -3,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "2",
                    3,
                    3,
                    0,
                    0,
                    1,
                    3,
                    1,
                    3,
                    1,
                    0,
                    0,
                    2.4,
                    0.8,
                    0.6,
                    0.2,
                    6,
                    2,
                    3,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    3,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
            ]

            for rowNumber in range(1, 4):
                values = allTimeOwnersValues[rowNumber - 1]
                for columnNumber, value in enumerate(values):
                    cell = f"{get_column_letter(columnNumber + 1)}{rowNumber}"
                    self.assertEqual(values[columnNumber], allTimeOwnersWorksheet[cell].value)
            # check that "next" cell is empty
            self.assertIsNone(allTimeOwnersWorksheet["A4"].value)
            # check legend
            self.assertEqual("Filters Applied", allTimeOwnersWorksheet["A6"].value)
            self.assertEqual("Week Number Start: 1", allTimeOwnersWorksheet["A7"].value)
            self.assertEqual("Year Number Start: 2000", allTimeOwnersWorksheet["A8"].value)
            self.assertEqual("Week Number End: 1", allTimeOwnersWorksheet["A9"].value)
            self.assertEqual("Year Number End: 2002", allTimeOwnersWorksheet["A10"].value)
            self.assertEqual("Only Regular Season: False", allTimeOwnersWorksheet["A11"].value)
            self.assertEqual("Only Post Season: False", allTimeOwnersWorksheet["A12"].value)
            self.assertEqual("Only Championship: False", allTimeOwnersWorksheet["A13"].value)
            self.assertIsNone(allTimeOwnersWorksheet["A14"].value)

    def test_leagueToExcel_filePathNotGiven_happyPath(self):
        owners, teams1 = getNDefaultOwnersAndTeams(2)
        teams1[0].name = "a"
        teams1[1].name = "b"
        matchup1 = Matchup(teamAId=teams1[0].id, teamBId=teams1[1].id, teamAScore=1, teamBScore=2)
        week1 = Week(weekNumber=1, matchups=[matchup1])
        year1 = Year(yearNumber=2000, teams=teams1, weeks=[week1])

        teams2 = [Team(ownerId=owners[0].id, name="a2"), Team(ownerId=owners[1].id, name="b2")]
        matchup2 = Matchup(teamAId=teams2[0].id, teamBId=teams2[1].id, teamAScore=1, teamBScore=2)
        week2 = Week(weekNumber=1, matchups=[matchup2])
        year2 = Year(yearNumber=2001, teams=teams2, weeks=[week2])

        teams3 = [Team(ownerId=owners[0].id, name="a3"), Team(ownerId=owners[1].id, name="b3")]
        matchup3 = Matchup(teamAId=teams3[0].id, teamBId=teams3[1].id, teamAScore=1, teamBScore=2)
        week3 = Week(weekNumber=1, matchups=[matchup3])
        year3 = Year(yearNumber=2002, teams=teams3, weeks=[week3])

        league = League(name="", owners=owners, years=[year1, year2, year3])

        with tempfile.TemporaryDirectory() as tempDir:
            fullPath = os.path.join(tempDir, "tmp.xlsx")
            leagueToExcel(league, fullPath)

            # open created Excel file to check that it was saved correctly
            workbook = load_workbook(filename=fullPath)

            # check sheets
            self.assertEqual(9, len(workbook.sheetnames))
            self.assertEqual("2000 Teams", workbook.sheetnames[0])
            self.assertEqual("2000 Matchups", workbook.sheetnames[1])
            self.assertEqual("2001 Teams", workbook.sheetnames[2])
            self.assertEqual("2001 Matchups", workbook.sheetnames[3])
            self.assertEqual("2002 Teams", workbook.sheetnames[4])
            self.assertEqual("2002 Matchups", workbook.sheetnames[5])
            self.assertEqual("All Time Teams", workbook.sheetnames[6])
            self.assertEqual("All Time Matchups", workbook.sheetnames[7])
            self.assertEqual("All Time Owners", workbook.sheetnames[8])

            # check sheet values
            # test year worksheet values
            worksheet2000TeamValues = [
                [
                    "Team",
                    "Games Played",
                    "Wins",
                    "Losses",
                    "Ties",
                    "Win Percentage",
                    "WAL",
                    "WAL Per Game",
                    "AWAL",
                    "AWAL Per Game",
                    "Opponent AWAL",
                    "Opponent AWAL Per Game",
                    "Smart Wins",
                    "Smart Wins Per Game",
                    "Opponent Smart Wins",
                    "Opponent Smart Wins Per Game",
                    "Points Scored",
                    "Points Scored Per Game",
                    "Opponent Points Scored",
                    "Opponent Points Scored Per Game",
                    "Scoring Share",
                    "Opponent Scoring Share",
                    "Max Scoring Share",
                    "Min Scoring Share",
                    "Max Score",
                    "Min Score",
                    "Scoring Standard Deviation",
                    "Plus/Minus",
                    "Team Score",
                    "Team Success",
                    "Team Luck",
                ],
                [
                    "a",
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b",
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
            ]

            worksheet2000MatchupValues = [
                [
                    "Team For",
                    "Team Against",
                    "Week Number",
                    "Matchup Type",
                    "Points For",
                    "Points Against",
                ],
                ["a", "b", 1, "REGULAR_SEASON", 1, 2],
                ["b", "a", 1, "REGULAR_SEASON", 2, 1],
            ]

            worksheet2001TeamValues = [
                [
                    "Team",
                    "Games Played",
                    "Wins",
                    "Losses",
                    "Ties",
                    "Win Percentage",
                    "WAL",
                    "WAL Per Game",
                    "AWAL",
                    "AWAL Per Game",
                    "Opponent AWAL",
                    "Opponent AWAL Per Game",
                    "Smart Wins",
                    "Smart Wins Per Game",
                    "Opponent Smart Wins",
                    "Opponent Smart Wins Per Game",
                    "Points Scored",
                    "Points Scored Per Game",
                    "Opponent Points Scored",
                    "Opponent Points Scored Per Game",
                    "Scoring Share",
                    "Opponent Scoring Share",
                    "Max Scoring Share",
                    "Min Scoring Share",
                    "Max Score",
                    "Min Score",
                    "Scoring Standard Deviation",
                    "Plus/Minus",
                    "Team Score",
                    "Team Success",
                    "Team Luck",
                ],
                [
                    "a2",
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b2",
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
            ]

            worksheet2001MatchupValues = [
                [
                    "Team For",
                    "Team Against",
                    "Week Number",
                    "Matchup Type",
                    "Points For",
                    "Points Against",
                ],
                ["a2", "b2", 1, "REGULAR_SEASON", 1, 2],
                ["b2", "a2", 1, "REGULAR_SEASON", 2, 1],
            ]

            worksheet2002TeamValues = [
                [
                    "Team",
                    "Games Played",
                    "Wins",
                    "Losses",
                    "Ties",
                    "Win Percentage",
                    "WAL",
                    "WAL Per Game",
                    "AWAL",
                    "AWAL Per Game",
                    "Opponent AWAL",
                    "Opponent AWAL Per Game",
                    "Smart Wins",
                    "Smart Wins Per Game",
                    "Opponent Smart Wins",
                    "Opponent Smart Wins Per Game",
                    "Points Scored",
                    "Points Scored Per Game",
                    "Opponent Points Scored",
                    "Opponent Points Scored Per Game",
                    "Scoring Share",
                    "Opponent Scoring Share",
                    "Max Scoring Share",
                    "Min Scoring Share",
                    "Max Score",
                    "Min Score",
                    "Scoring Standard Deviation",
                    "Plus/Minus",
                    "Team Score",
                    "Team Success",
                    "Team Luck",
                ],
                [
                    "a3",
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b3",
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
            ]

            worksheet2002MatchupValues = [
                [
                    "Team For",
                    "Team Against",
                    "Week Number",
                    "Matchup Type",
                    "Points For",
                    "Points Against",
                ],
                ["a3", "b3", 1, "REGULAR_SEASON", 1, 2],
                ["b3", "a3", 1, "REGULAR_SEASON", 2, 1],
            ]

            yearWorksheetsAndValues = [
                (workbook["2000 Teams"], worksheet2000TeamValues),
                (workbook["2000 Matchups"], worksheet2000MatchupValues),
                (workbook["2001 Teams"], worksheet2001TeamValues),
                (workbook["2001 Matchups"], worksheet2001MatchupValues),
                (workbook["2002 Teams"], worksheet2002TeamValues),
                (workbook["2002 Matchups"], worksheet2002MatchupValues),
            ]

            for i, (currentWorksheet, allWorksheetValues) in enumerate(yearWorksheetsAndValues):
                for j, worksheetValues in enumerate(allWorksheetValues):
                    for k, worksheetValue in enumerate(worksheetValues):
                        row = j + 1
                        column = get_column_letter(k + 1)
                        cellValue = currentWorksheet[f"{column}{row}"].value
                        if isinstance(cellValue, Decimal):
                            cellValue = float(cellValue)
                        self.assertEqual(worksheetValue, cellValue)
                # check that "next" cell is empty
                self.assertIsNone(currentWorksheet["A4"].value)
                # check legend
                self.assertEqual("Filters Applied", currentWorksheet["A6"].value)
                self.assertEqual("Week Number Start: 1", currentWorksheet["A7"].value)
                self.assertEqual("Week Number End: 1", currentWorksheet["A8"].value)
                self.assertEqual("Only Regular Season: False", currentWorksheet["A9"].value)
                self.assertEqual("Only Post Season: False", currentWorksheet["A10"].value)
                self.assertEqual("Only Championship: False", currentWorksheet["A11"].value)
                self.assertEqual("Include Multi-Week Matchups: True", currentWorksheet["A12"].value)
                self.assertIsNone(currentWorksheet["A13"].value)

            # test all time teams worksheet values
            allTimeTeamsWorksheet = workbook["All Time Teams"]

            allTimeTeamsValues = [
                [
                    "Team",
                    "Owner",
                    "Year",
                    "Games Played",
                    "Wins",
                    "Losses",
                    "Ties",
                    "Win Percentage",
                    "WAL",
                    "WAL Per Game",
                    "AWAL",
                    "AWAL Per Game",
                    "Opponent AWAL",
                    "Opponent AWAL Per Game",
                    "Smart Wins",
                    "Smart Wins Per Game",
                    "Opponent Smart Wins",
                    "Opponent Smart Wins Per Game",
                    "Points Scored",
                    "Points Scored Per Game",
                    "Opponent Points Scored",
                    "Opponent Points Scored Per Game",
                    "Scoring Share",
                    "Opponent Scoring Share",
                    "Max Scoring Share",
                    "Min Scoring Share",
                    "Max Score",
                    "Min Score",
                    "Scoring Standard Deviation",
                    "Plus/Minus",
                    "Team Score",
                    "Team Success",
                    "Team Luck",
                ],
                [
                    "a",
                    "1",
                    2000,
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b",
                    "2",
                    2000,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
                [
                    "a2",
                    "1",
                    2001,
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b2",
                    "2",
                    2001,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
                [
                    "a3",
                    "1",
                    2002,
                    1,
                    0,
                    1,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    2,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -1,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "b3",
                    "2",
                    2002,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    1,
                    1,
                    1,
                    0,
                    0,
                    1,
                    1,
                    0,
                    0,
                    2,
                    2,
                    1,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    1,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
            ]

            for rowNumber in range(1, 8):
                values = allTimeTeamsValues[rowNumber - 1]
                for columnNumber, value in enumerate(values):
                    cell = f"{get_column_letter(columnNumber + 1)}{rowNumber}"
                    cellValue = allTimeTeamsWorksheet[cell].value
                    if isinstance(cellValue, Decimal):
                        cellValue = float(cellValue)
                    self.assertEqual(values[columnNumber], cellValue)
            # check that "next" cell is empty
            self.assertIsNone(allTimeTeamsWorksheet["A8"].value)
            # check legend
            self.assertEqual("Filters Applied", allTimeTeamsWorksheet["A10"].value)
            self.assertEqual("Week Number Start: 1", allTimeTeamsWorksheet["A11"].value)
            self.assertEqual("Year Number Start: 2000", allTimeTeamsWorksheet["A12"].value)
            self.assertEqual("Week Number End: 1", allTimeTeamsWorksheet["A13"].value)
            self.assertEqual("Year Number End: 2002", allTimeTeamsWorksheet["A14"].value)
            self.assertEqual("Only Regular Season: False", allTimeTeamsWorksheet["A15"].value)
            self.assertEqual("Only Post Season: False", allTimeTeamsWorksheet["A16"].value)
            self.assertEqual("Only Championship: False", allTimeTeamsWorksheet["A17"].value)
            self.assertIsNone(allTimeTeamsWorksheet["A18"].value)

            # test all time matchups worksheet values
            allTimeMatchupsWorksheet = workbook["All Time Matchups"]

            allTimeMatchupsValues = [
                [
                    "Team For",
                    "Owner For",
                    "Team Against",
                    "Owner Against",
                    "Year",
                    "Week Number",
                    "Matchup Type",
                    "Points For",
                    "Points Against",
                ],
                ["a", "1", "b", "2", 2000, 1, "REGULAR_SEASON", 1, 2],
                ["b", "2", "a", "1", 2000, 1, "REGULAR_SEASON", 2, 1],
                ["a2", "1", "b2", "2", 2001, 1, "REGULAR_SEASON", 1, 2],
                ["b2", "2", "a2", "1", 2001, 1, "REGULAR_SEASON", 2, 1],
                ["a3", "1", "b3", "2", 2002, 1, "REGULAR_SEASON", 1, 2],
                ["b3", "2", "a3", "1", 2002, 1, "REGULAR_SEASON", 2, 1],
            ]

            for rowNumber in range(1, 8):
                values = allTimeMatchupsValues[rowNumber - 1]
                for columnNumber, value in enumerate(values):
                    cell = f"{get_column_letter(columnNumber + 1)}{rowNumber}"
                    cellValue = allTimeMatchupsWorksheet[cell].value
                    if isinstance(cellValue, Decimal):
                        cellValue = float(cellValue)
                    self.assertEqual(values[columnNumber], cellValue)
            # check that "next" cell is empty
            self.assertIsNone(allTimeMatchupsWorksheet["A8"].value)
            # check legend
            self.assertEqual("Filters Applied", allTimeMatchupsWorksheet["A10"].value)
            self.assertEqual("Week Number Start: 1", allTimeMatchupsWorksheet["A11"].value)
            self.assertEqual("Year Number Start: 2000", allTimeMatchupsWorksheet["A12"].value)
            self.assertEqual("Week Number End: 1", allTimeMatchupsWorksheet["A13"].value)
            self.assertEqual("Year Number End: 2002", allTimeMatchupsWorksheet["A14"].value)
            self.assertEqual("Only Regular Season: False", allTimeMatchupsWorksheet["A15"].value)
            self.assertEqual("Only Post Season: False", allTimeMatchupsWorksheet["A16"].value)
            self.assertEqual("Only Championship: False", allTimeMatchupsWorksheet["A17"].value)
            self.assertIsNone(allTimeMatchupsWorksheet["A18"].value)

            # test all time owners worksheet values
            allTimeOwnersWorksheet = workbook["All Time Owners"]

            allTimeOwnersValues = [
                [
                    "Owner",
                    "Games Played",
                    "Wins",
                    "Losses",
                    "Ties",
                    "Win Percentage",
                    "WAL",
                    "WAL Per Game",
                    "AWAL",
                    "AWAL Per Game",
                    "Opponent AWAL",
                    "Opponent AWAL Per Game",
                    "Smart Wins",
                    "Smart Wins Per Game",
                    "Opponent Smart Wins",
                    "Opponent Smart Wins Per Game",
                    "Points Scored",
                    "Points Scored Per Game",
                    "Opponent Points Scored",
                    "Opponent Points Scored Per Game",
                    "Scoring Share",
                    "Opponent Scoring Share",
                    "Max Scoring Share",
                    "Min Scoring Share",
                    "Max Score",
                    "Min Score",
                    "Scoring Standard Deviation",
                    "Plus/Minus",
                    "Adjusted Team Score",
                    "Adjusted Team Success",
                    "Adjusted Team Luck",
                ],
                [
                    "1",
                    3,
                    0,
                    3,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    3,
                    1,
                    0.6,
                    0.2,
                    2.4,
                    0.8,
                    3,
                    1,
                    6,
                    2,
                    33.33333333333334,
                    66.66666666666667,
                    33.33333333333334,
                    33.33333333333334,
                    1,
                    1,
                    0,
                    -3,
                    66.76666666666667,
                    66.76666666666667,
                    0,
                ],
                [
                    "2",
                    3,
                    3,
                    0,
                    0,
                    1,
                    3,
                    1,
                    3,
                    1,
                    0,
                    0,
                    2.4,
                    0.8,
                    0.6,
                    0.2,
                    6,
                    2,
                    3,
                    1,
                    66.66666666666667,
                    33.33333333333334,
                    66.66666666666667,
                    66.66666666666667,
                    2,
                    2,
                    0,
                    3,
                    233.5333333333333,
                    233.5333333333333,
                    0,
                ],
            ]

            for rowNumber in range(1, 4):
                values = allTimeOwnersValues[rowNumber - 1]
                for columnNumber, value in enumerate(values):
                    cell = f"{get_column_letter(columnNumber + 1)}{rowNumber}"
                    cellValue = allTimeOwnersWorksheet[cell].value
                    if isinstance(cellValue, Decimal):
                        cellValue = float(cellValue)
                    self.assertEqual(values[columnNumber], cellValue)
            # check that "next" cell is empty
            self.assertIsNone(allTimeOwnersWorksheet["A4"].value)
            # check legend
            self.assertEqual("Filters Applied", allTimeOwnersWorksheet["A6"].value)
            self.assertEqual("Week Number Start: 1", allTimeOwnersWorksheet["A7"].value)
            self.assertEqual("Year Number Start: 2000", allTimeOwnersWorksheet["A8"].value)
            self.assertEqual("Week Number End: 1", allTimeOwnersWorksheet["A9"].value)
            self.assertEqual("Year Number End: 2002", allTimeOwnersWorksheet["A10"].value)
            self.assertEqual("Only Regular Season: False", allTimeOwnersWorksheet["A11"].value)
            self.assertEqual("Only Post Season: False", allTimeOwnersWorksheet["A12"].value)
            self.assertEqual("Only Championship: False", allTimeOwnersWorksheet["A13"].value)
            self.assertIsNone(allTimeOwnersWorksheet["A14"].value)

    def test_leagueToExcel_returnedWorkbookIsSameAsLoadedWorkbook(self):
        owners, teams1 = getNDefaultOwnersAndTeams(2)
        teams1[0].name = "a"
        teams1[1].name = "b"
        matchup1 = Matchup(teamAId=teams1[0].id, teamBId=teams1[1].id, teamAScore=1, teamBScore=2)
        week1 = Week(weekNumber=1, matchups=[matchup1])
        year1 = Year(yearNumber=2000, teams=teams1, weeks=[week1])

        teams2 = [Team(ownerId=owners[0].id, name="a2"), Team(ownerId=owners[1].id, name="b2")]
        matchup2 = Matchup(teamAId=teams2[0].id, teamBId=teams2[1].id, teamAScore=1, teamBScore=2)
        week2 = Week(weekNumber=1, matchups=[matchup2])
        year2 = Year(yearNumber=2001, teams=teams2, weeks=[week2])

        teams3 = [Team(ownerId=owners[0].id, name="a3"), Team(ownerId=owners[1].id, name="b3")]
        matchup3 = Matchup(teamAId=teams3[0].id, teamBId=teams3[1].id, teamAScore=1, teamBScore=2)
        week3 = Week(weekNumber=1, matchups=[matchup3])
        year3 = Year(yearNumber=2002, teams=teams3, weeks=[week3])

        league = League(name="", owners=owners, years=[year1, year2, year3])

        with tempfile.TemporaryDirectory() as tempDir:
            fullPath = os.path.join(tempDir, "tmp.xlsx")
            returnedWorkbook = leagueToExcel(league, fullPath)

            # open created Excel file to check that it was saved correctly
            loadedWorkbook = load_workbook(filename=fullPath)

            self.assertEqual(len(returnedWorkbook.worksheets), len(loadedWorkbook.worksheets))
            for ws1, ws2 in zip(returnedWorkbook.worksheets, loadedWorkbook.worksheets):
                # Compare worksheet names and contents
                self.assertEqual(ws1.title, ws2.title)
                # Compare cells
                for row1, row2 in zip(ws1.rows, ws2.rows):
                    for cell1, cell2 in zip(row1, row2):
                        cell1Value = cell1.value
                        cell2Value = cell2.value
                        if isinstance(cell1Value, Decimal):
                            cell1Value = float(cell1Value)
                        if isinstance(cell2Value, Decimal):
                            cell2Value = float(cell2Value)
                        self.assertAlmostEqual(cell1Value, cell2Value)

    def test_leagueToExcel_leagueIsNone_raisesException(self):

        with self.assertRaises(ValueError) as context:
            leagueToExcel(None, "")
        self.assertEqual("'league' has not been set.", str(context.exception))

    def test_leagueToExcel_fileAlreadyExists_raisesException(self):
        owners, teams1 = getNDefaultOwnersAndTeams(2)
        teams1[0].name = "a"
        teams1[1].name = "b"
        matchup1 = Matchup(teamAId=teams1[0].id, teamBId=teams1[1].id, teamAScore=1, teamBScore=2)
        week1 = Week(weekNumber=1, matchups=[matchup1])
        year1 = Year(yearNumber=2000, teams=teams1, weeks=[week1])

        teams2 = [Team(ownerId=owners[0].id, name="a2"), Team(ownerId=owners[1].id, name="b2")]
        matchup2 = Matchup(teamAId=teams2[0].id, teamBId=teams2[1].id, teamAScore=1, teamBScore=2)
        week2 = Week(weekNumber=1, matchups=[matchup2])
        year2 = Year(yearNumber=2001, teams=teams2, weeks=[week2])

        teams3 = [Team(ownerId=owners[0].id, name="a3"), Team(ownerId=owners[1].id, name="b3")]
        matchup3 = Matchup(teamAId=teams3[0].id, teamBId=teams3[1].id, teamAScore=1, teamBScore=2)
        week3 = Week(weekNumber=1, matchups=[matchup3])
        year3 = Year(yearNumber=1999, teams=teams3, weeks=[week3])

        league = League(name="", owners=owners, years=[year1, year2, year3])

        with self.assertRaises(FileExistsError) as context:
            with tempfile.TemporaryDirectory() as tempDir:
                fullPath = os.path.join(tempDir, "tmp.xlsx")
                # create a file with this same name/path first
                with open(fullPath, "w") as f:
                    ...
                leagueToExcel(league, fullPath)
        self.assertEqual(
            f"Cannot create file at path: '{fullPath}' because there is already a file there.",
            str(context.exception),
        )
