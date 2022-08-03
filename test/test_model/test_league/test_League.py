import os
import tempfile
import unittest

from openpyxl import load_workbook

from leeger.enum.MatchupType import MatchupType
from leeger.exception.InvalidLeagueFormatException import InvalidLeagueFormatException
from leeger.model.league.League import League
from leeger.model.league.Matchup import Matchup
from leeger.model.league.Owner import Owner
from leeger.model.league.Team import Team
from leeger.model.league.Week import Week
from leeger.model.league.Year import Year
from leeger.model.stat.AllTimeStatSheet import AllTimeStatSheet
from test.helper.prototypes import getNDefaultOwnersAndTeams


class TestLeague(unittest.TestCase):
    def test_league_init(self):
        year = Year(yearNumber=0, teams=[], weeks=[])
        owner = Owner(name="")
        league = League(name="leagueName", owners=[owner], years=[year])

        self.assertEqual("leagueName", league.name)
        self.assertEqual(1, len(league.owners))
        self.assertEqual(1, len(league.years))
        self.assertEqual(owner.id, league.owners[0].id)
        self.assertEqual(year.id, league.years[0].id)

    def test_league_add_happyPath(self):
        # create League 1
        owners_1, teams_1 = getNDefaultOwnersAndTeams(2)

        matchup_1 = Matchup(teamAId=teams_1[0].id, teamBId=teams_1[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_1 = Week(weekNumber=1, matchups=[matchup_1])
        year_1 = Year(yearNumber=2000, teams=teams_1, weeks=[week_1])
        league_1 = League(name="LEAGUE 1", owners=owners_1, years=[year_1])

        # create League 2
        owners_2, teams_2 = getNDefaultOwnersAndTeams(2)

        matchup_2 = Matchup(teamAId=teams_2[0].id, teamBId=teams_2[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_2 = Week(weekNumber=1, matchups=[matchup_2])
        year_2 = Year(yearNumber=2001, teams=teams_2, weeks=[week_2])
        league_2 = League(name="LEAGUE 2", owners=owners_2, years=[year_2])

        combinedLeague = league_1 + league_2
        self.assertIsInstance(combinedLeague, League)
        self.assertEqual("'LEAGUE 1' + 'LEAGUE 2' League", combinedLeague.name)
        self.assertEqual(2, len(combinedLeague.owners))
        self.assertEqual([owner.name for owner in league_1.owners], [owner.name for owner in league_1.owners])
        self.assertEqual(2, len(combinedLeague.years))
        self.assertEqual(2000, combinedLeague.years[0].yearNumber)
        self.assertEqual(2001, combinedLeague.years[1].yearNumber)

    def test_league_add_addingOrderDoesNotMatter(self):
        # create League 1
        owners_1, teams_1 = getNDefaultOwnersAndTeams(2)

        matchup_1 = Matchup(teamAId=teams_1[0].id, teamBId=teams_1[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_1 = Week(weekNumber=1, matchups=[matchup_1])
        year_1 = Year(yearNumber=2000, teams=teams_1, weeks=[week_1])
        league_1 = League(name="LEAGUE", owners=owners_1, years=[year_1])

        # create League 2
        owners_2, teams_2 = getNDefaultOwnersAndTeams(2)

        matchup_2 = Matchup(teamAId=teams_2[0].id, teamBId=teams_2[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_2 = Week(weekNumber=1, matchups=[matchup_2])
        year_2 = Year(yearNumber=2001, teams=teams_2, weeks=[week_2])
        league_2 = League(name="LEAGUE", owners=owners_2, years=[year_2])

        combinedLeague1 = league_1 + league_2
        combinedLeague2 = league_2 + league_1
        self.assertIsInstance(combinedLeague1, League)
        self.assertIsInstance(combinedLeague2, League)
        self.assertEqual(combinedLeague1, combinedLeague2)

    def test_league_add_sameLeagueName_nameIsntChanged(self):
        # create League 1
        owners_1, teams_1 = getNDefaultOwnersAndTeams(2)

        matchup_1 = Matchup(teamAId=teams_1[0].id, teamBId=teams_1[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_1 = Week(weekNumber=1, matchups=[matchup_1])
        year_1 = Year(yearNumber=2000, teams=teams_1, weeks=[week_1])
        league_1 = League(name="LEAGUE", owners=owners_1, years=[year_1])

        # create League 2
        owners_2, teams_2 = getNDefaultOwnersAndTeams(2)

        matchup_2 = Matchup(teamAId=teams_2[0].id, teamBId=teams_2[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_2 = Week(weekNumber=1, matchups=[matchup_2])
        year_2 = Year(yearNumber=2001, teams=teams_2, weeks=[week_2])
        league_2 = League(name="LEAGUE", owners=owners_2, years=[year_2])

        combinedLeague = league_1 + league_2
        self.assertIsInstance(combinedLeague, League)
        self.assertEqual("LEAGUE", combinedLeague.name)

    def test_league_add_leaguesHaveOwnersThatDontHaveMatchingNames_ownersAreCombined(self):
        # create League 1
        owners_1, teams_1 = getNDefaultOwnersAndTeams(2)
        owners_1[0].name = "league 1 owner 1"
        owners_1[1].name = "league 1 owner 2"

        matchup_1 = Matchup(teamAId=teams_1[0].id, teamBId=teams_1[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_1 = Week(weekNumber=1, matchups=[matchup_1])
        year_1 = Year(yearNumber=2000, teams=teams_1, weeks=[week_1])
        league_1 = League(name="LEAGUE", owners=owners_1, years=[year_1])

        # create League 2
        owners_2, teams_2 = getNDefaultOwnersAndTeams(2)
        owners_2[0].name = "league 2 owner 1"
        owners_2[1].name = "league 2 owner 2"

        matchup_2 = Matchup(teamAId=teams_2[0].id, teamBId=teams_2[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_2 = Week(weekNumber=1, matchups=[matchup_2])
        year_2 = Year(yearNumber=2001, teams=teams_2, weeks=[week_2])
        league_2 = League(name="LEAGUE", owners=owners_2, years=[year_2])

        combinedLeague = league_1 + league_2
        self.assertIsInstance(combinedLeague, League)
        self.assertEqual(4, len(combinedLeague.owners))
        self.assertEqual("league 1 owner 1", combinedLeague.owners[0].name)
        self.assertEqual("league 1 owner 2", combinedLeague.owners[1].name)
        self.assertEqual("league 2 owner 1", combinedLeague.owners[2].name)
        self.assertEqual("league 2 owner 2", combinedLeague.owners[3].name)

    def test_league_add_combinedLeagueYearsAreInCorrectOrder(self):
        # first example, League 1 has years that come first
        # create League 1
        owners_1, teams_1 = getNDefaultOwnersAndTeams(2)

        matchup_1 = Matchup(teamAId=teams_1[0].id, teamBId=teams_1[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_1 = Week(weekNumber=1, matchups=[matchup_1])
        year_1 = Year(yearNumber=2000, teams=teams_1, weeks=[week_1])
        league_1 = League(name="LEAGUE 1", owners=owners_1, years=[year_1])

        # create League 2
        owners_2, teams_2 = getNDefaultOwnersAndTeams(2)

        matchup_2 = Matchup(teamAId=teams_2[0].id, teamBId=teams_2[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_2 = Week(weekNumber=1, matchups=[matchup_2])
        year_2 = Year(yearNumber=2001, teams=teams_2, weeks=[week_2])
        league_2 = League(name="LEAGUE 2", owners=owners_2, years=[year_2])

        combinedLeague = league_1 + league_2
        self.assertIsInstance(combinedLeague, League)
        self.assertEqual(2000, combinedLeague.years[0].yearNumber)
        self.assertEqual(2001, combinedLeague.years[1].yearNumber)

        # second example, League 2 has years that come first
        # create League 1
        owners_1, teams_1 = getNDefaultOwnersAndTeams(2)

        matchup_1 = Matchup(teamAId=teams_1[0].id, teamBId=teams_1[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_1 = Week(weekNumber=1, matchups=[matchup_1])
        year_1 = Year(yearNumber=2001, teams=teams_1, weeks=[week_1])
        league_1 = League(name="LEAGUE 1", owners=owners_1, years=[year_1])

        # create League 2
        owners_2, teams_2 = getNDefaultOwnersAndTeams(2)

        matchup_2 = Matchup(teamAId=teams_2[0].id, teamBId=teams_2[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_2 = Week(weekNumber=1, matchups=[matchup_2])
        year_2 = Year(yearNumber=2000, teams=teams_2, weeks=[week_2])
        league_2 = League(name="LEAGUE 2", owners=owners_2, years=[year_2])

        combinedLeague = league_1 + league_2
        self.assertIsInstance(combinedLeague, League)
        self.assertEqual(2000, combinedLeague.years[0].yearNumber)
        self.assertEqual(2001, combinedLeague.years[1].yearNumber)

    def test_league_add_duplicateYearsAcrossCombinedLeagues_raisesException(self):
        # create League 1
        owners_1, teams_1 = getNDefaultOwnersAndTeams(2)

        matchup_1 = Matchup(teamAId=teams_1[0].id, teamBId=teams_1[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_1 = Week(weekNumber=1, matchups=[matchup_1])
        year_1 = Year(yearNumber=2000, teams=teams_1, weeks=[week_1])
        league_1 = League(name="LEAGUE 1", owners=owners_1, years=[year_1])

        # create League 2
        owners_2, teams_2 = getNDefaultOwnersAndTeams(2)

        matchup_2 = Matchup(teamAId=teams_2[0].id, teamBId=teams_2[1].id, teamAScore=1.1, teamBScore=2.2,
                            matchupType=MatchupType.REGULAR_SEASON)
        week_2 = Week(weekNumber=1, matchups=[matchup_2])
        year_2 = Year(yearNumber=2000, teams=teams_2, weeks=[week_2])
        league_2 = League(name="LEAGUE 2", owners=owners_2, years=[year_2])

        with self.assertRaises(InvalidLeagueFormatException) as context:
            league_1 + league_2
        self.assertEqual("Can only have 1 of each year number within a league.", str(context.exception))

    def test_league_statSheet(self):
        owners, teams = getNDefaultOwnersAndTeams(2)

        matchup = Matchup(teamAId=teams[0].id, teamBId=teams[1].id, teamAScore=1, teamBScore=2)
        week = Week(weekNumber=1, matchups=[matchup])
        year = Year(yearNumber=2000, teams=teams, weeks=[week])

        league = League(name="TEST", owners=owners, years=[year])

        leagueStatSheet = league.statSheet()

        self.assertIsInstance(leagueStatSheet, AllTimeStatSheet)
        self.assertIsInstance(leagueStatSheet.wins, dict)
        self.assertIsInstance(leagueStatSheet.losses, dict)
        self.assertIsInstance(leagueStatSheet.ties, dict)
        self.assertIsInstance(leagueStatSheet.winPercentage, dict)
        self.assertIsInstance(leagueStatSheet.wal, dict)
        self.assertIsInstance(leagueStatSheet.walPerGame, dict)

        self.assertIsInstance(leagueStatSheet.awal, dict)
        self.assertIsInstance(leagueStatSheet.awalPerGame, dict)
        self.assertIsInstance(leagueStatSheet.opponentAWAL, dict)
        self.assertIsInstance(leagueStatSheet.opponentAWALPerGame, dict)

        self.assertIsInstance(leagueStatSheet.smartWins, dict)
        self.assertIsInstance(leagueStatSheet.smartWinsPerGame, dict)
        self.assertIsInstance(leagueStatSheet.opponentSmartWins, dict)
        self.assertIsInstance(leagueStatSheet.opponentSmartWinsPerGame, dict)

        self.assertIsInstance(leagueStatSheet.pointsScored, dict)
        self.assertIsInstance(leagueStatSheet.pointsScoredPerGame, dict)
        self.assertIsInstance(leagueStatSheet.opponentPointsScored, dict)
        self.assertIsInstance(leagueStatSheet.opponentPointsScoredPerGame, dict)

        self.assertIsInstance(leagueStatSheet.scoringShare, dict)
        self.assertIsInstance(leagueStatSheet.opponentScoringShare, dict)

        self.assertIsInstance(leagueStatSheet.maxScore, dict)
        self.assertIsInstance(leagueStatSheet.minScore, dict)

        self.assertIsInstance(leagueStatSheet.scoringStandardDeviation, dict)

        self.assertIsInstance(leagueStatSheet.plusMinus, dict)

        self.assertIsInstance(leagueStatSheet.teamScore, dict)
        self.assertIsInstance(leagueStatSheet.teamSuccess, dict)
        self.assertIsInstance(leagueStatSheet.teamLuck, dict)

    def test_league_toExcel_happyPath(self):
        owners, teams1 = getNDefaultOwnersAndTeams(2)
        teams1[0].name = "a"
        teams1[1].name = "b"
        matchup1 = Matchup(teamAId=teams1[0].id, teamBId=teams1[1].id, teamAScore=1, teamBScore=2)
        week1 = Week(weekNumber=1, matchups=[matchup1])
        year1 = Year(yearNumber=2000, teams=teams1, weeks=[week1])

        teams2 = [Team(ownerId=owners[0].id, name="a2"), Team(ownerId=owners[1].id, name="b2")]
        matchup2 = Matchup(teamAId=teams2[0].id, teamBId=teams2[1].id, teamAScore=1, teamBScore=2)
        week2 = Week(weekNumber=1, matchups=[matchup2])
        year2 = Year(yearNumber=2001, teams=teams2, weeks=[week2])

        teams3 = [Team(ownerId=owners[0].id, name="a3"), Team(ownerId=owners[1].id, name="b3")]
        matchup3 = Matchup(teamAId=teams3[0].id, teamBId=teams3[1].id, teamAScore=1, teamBScore=2)
        week3 = Week(weekNumber=1, matchups=[matchup3])
        year3 = Year(yearNumber=2002, teams=teams3, weeks=[week3])

        league = League(name="", owners=owners, years=[year1, year2, year3])

        with tempfile.TemporaryDirectory() as tempDir:
            fullPath = os.path.join(tempDir, "tmp.xlsx")
            league.toExcel(fullPath)

            # open created Excel file to check that it was saved correctly
            workbook = load_workbook(filename=fullPath)

            self.assertEqual(4, len(workbook.sheetnames))
            self.assertEqual("2000", workbook.sheetnames[0])
            self.assertEqual("2001", workbook.sheetnames[1])
            self.assertEqual("2002", workbook.sheetnames[2])
            self.assertEqual("All Time", workbook.sheetnames[3])

    def test_league_toExcel_fileAlreadyExists_raisesException(self):
        owners, teams1 = getNDefaultOwnersAndTeams(2)
        teams1[0].name = "a"
        teams1[1].name = "b"
        matchup1 = Matchup(teamAId=teams1[0].id, teamBId=teams1[1].id, teamAScore=1, teamBScore=2)
        week1 = Week(weekNumber=1, matchups=[matchup1])
        year1 = Year(yearNumber=2000, teams=teams1, weeks=[week1])

        teams2 = [Team(ownerId=owners[0].id, name="a2"), Team(ownerId=owners[1].id, name="b2")]
        matchup2 = Matchup(teamAId=teams2[0].id, teamBId=teams2[1].id, teamAScore=1, teamBScore=2)
        week2 = Week(weekNumber=1, matchups=[matchup2])
        year2 = Year(yearNumber=2001, teams=teams2, weeks=[week2])

        teams3 = [Team(ownerId=owners[0].id, name="a3"), Team(ownerId=owners[1].id, name="b3")]
        matchup3 = Matchup(teamAId=teams3[0].id, teamBId=teams3[1].id, teamAScore=1, teamBScore=2)
        week3 = Week(weekNumber=1, matchups=[matchup3])
        year3 = Year(yearNumber=1999, teams=teams3, weeks=[week3])

        league = League(name="", owners=owners, years=[year1, year2, year3])

        with self.assertRaises(FileExistsError) as context:
            with tempfile.TemporaryDirectory() as tempDir:
                fullPath = os.path.join(tempDir, "tmp.xlsx")
                # create a file with this same name/path first
                with open(fullPath, "w") as f:
                    ...
                league.toExcel(fullPath)
        self.assertEqual(f"Cannot create file at path: '{fullPath}' because there is already a file there.",
                         str(context.exception))
