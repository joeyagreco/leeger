import math
import os
import tempfile
import unittest

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.leeger.model.league.Matchup import Matchup
from src.leeger.model.league.Team import Team
from src.leeger.model.league.Week import Week
from src.leeger.model.league.Year import Year
from src.leeger.model.stat.YearStatSheet import YearStatSheet
from test.helper.prototypes import getNDefaultOwnersAndTeams


class TestYear(unittest.TestCase):
    def test_year_init(self):
        week = Week(weekNumber=1, matchups=[])
        team = Team(ownerId="", name="")
        year = Year(yearNumber=2000, teams=[team], weeks=[week])

        self.assertEqual(2000, year.yearNumber)
        self.assertEqual(1, len(year.teams))
        self.assertEqual(1, len(year.weeks))
        self.assertEqual(week.id, year.weeks[0].id)
        self.assertEqual(team.id, year.teams[0].id)

    def test_year_statSheet(self):
        owners, teams = getNDefaultOwnersAndTeams(2)

        matchup = Matchup(teamAId=teams[0].id, teamBId=teams[1].id, teamAScore=1, teamBScore=2)
        week = Week(weekNumber=1, matchups=[matchup])
        year = Year(yearNumber=2000, teams=teams, weeks=[week])

        yearStatSheet = year.statSheet()

        self.assertIsInstance(yearStatSheet, YearStatSheet)
        self.assertIsInstance(yearStatSheet.wins, dict)
        self.assertIsInstance(yearStatSheet.losses, dict)
        self.assertIsInstance(yearStatSheet.ties, dict)
        self.assertIsInstance(yearStatSheet.winPercentage, dict)
        self.assertIsInstance(yearStatSheet.wal, dict)
        self.assertIsInstance(yearStatSheet.walPerGame, dict)

        self.assertIsInstance(yearStatSheet.awal, dict)
        self.assertIsInstance(yearStatSheet.awalPerGame, dict)
        self.assertIsInstance(yearStatSheet.opponentAWAL, dict)
        self.assertIsInstance(yearStatSheet.opponentAWALPerGame, dict)

        self.assertIsInstance(yearStatSheet.smartWins, dict)
        self.assertIsInstance(yearStatSheet.smartWinsPerGame, dict)
        self.assertIsInstance(yearStatSheet.opponentSmartWins, dict)
        self.assertIsInstance(yearStatSheet.opponentSmartWinsPerGame, dict)

        self.assertIsInstance(yearStatSheet.pointsScored, dict)
        self.assertIsInstance(yearStatSheet.pointsScoredPerGame, dict)
        self.assertIsInstance(yearStatSheet.opponentPointsScored, dict)
        self.assertIsInstance(yearStatSheet.opponentPointsScoredPerGame, dict)

        self.assertIsInstance(yearStatSheet.scoringShare, dict)
        self.assertIsInstance(yearStatSheet.opponentScoringShare, dict)

        self.assertIsInstance(yearStatSheet.maxScore, dict)
        self.assertIsInstance(yearStatSheet.minScore, dict)

        self.assertIsInstance(yearStatSheet.scoringStandardDeviation, dict)

        self.assertIsInstance(yearStatSheet.plusMinus, dict)

        self.assertIsInstance(yearStatSheet.teamScore, dict)
        self.assertIsInstance(yearStatSheet.teamSuccess, dict)
        self.assertIsInstance(yearStatSheet.teamLuck, dict)

        self.assertIsInstance(yearStatSheet.championshipCount, dict)

    def test_year_toExcel_excelSheetDoesntAlreadyExist(self):
        owners, teams = getNDefaultOwnersAndTeams(2)
        teams[0].name = "a"
        teams[1].name = "b"
        matchup = Matchup(teamAId=teams[0].id, teamBId=teams[1].id, teamAScore=1, teamBScore=2)
        week = Week(weekNumber=1, matchups=[matchup])
        year = Year(yearNumber=2000, teams=teams, weeks=[week])

        with tempfile.TemporaryDirectory() as tempDir:
            fullPath = os.path.join(tempDir, "tmp.xlsx")
            year.toExcel(fullPath)

            # open created Excel file to check that it was saved correctly
            workbook = load_workbook(filename=fullPath)
            worksheet = workbook.active

            self.assertEqual(1, len(workbook.sheetnames))
            self.assertEqual("2000", workbook.sheetnames[0])
            self.assertEqual("Team Names", worksheet["A1"].value)
            self.assertEqual("a", worksheet["A2"].value)
            self.assertEqual("b", worksheet["A3"].value)

            statsWithTitles = year.statSheet().preferredOrderWithTitle()
            for row, teamId in enumerate([team.id for team in year.teams]):
                for col, statWithTitle in enumerate(statsWithTitles):
                    char = get_column_letter(col + 2)
                    if row == 1:
                        # check stat header
                        self.assertEqual(statWithTitle[0], worksheet[f"{char}{row}"].value)
                    # check stat value
                    # due to Excel rounding values, we assert that the values are very, very close
                    assert math.isclose(float(statWithTitle[1][teamId]), worksheet[f"{char}{row + 2}"].value,
                                        rel_tol=0.000000000000001)

    def test_year_toExcel_excelSheetAlreadyExists(self):
        owners, teams1 = getNDefaultOwnersAndTeams(2)
        teams1[0].name = "a1"
        teams1[1].name = "b1"
        matchup1 = Matchup(teamAId=teams1[0].id, teamBId=teams1[1].id, teamAScore=1, teamBScore=2)
        week1 = Week(weekNumber=1, matchups=[matchup1])
        year1 = Year(yearNumber=2000, teams=teams1, weeks=[week1])

        teams2 = [Team(ownerId=owners[0].id, name="a2"), Team(ownerId=owners[1].id, name="b2")]
        matchup2 = Matchup(teamAId=teams2[0].id, teamBId=teams2[1].id, teamAScore=1, teamBScore=2)
        week2 = Week(weekNumber=1, matchups=[matchup2])
        year2 = Year(yearNumber=2001, teams=teams2, weeks=[week2])

        with tempfile.TemporaryDirectory() as tempDir:
            fullPath = os.path.join(tempDir, "tmp.xlsx")
            year1.toExcel(fullPath)

            year2.toExcel(fullPath)

            # open created Excel file to check that it was saved correctly
            workbook = load_workbook(filename=fullPath)
            worksheet = workbook.worksheets[workbook.sheetnames.index("2001")]

            self.assertEqual(2, len(workbook.sheetnames))
            self.assertEqual("2000", workbook.sheetnames[0])
            self.assertEqual("2001", workbook.sheetnames[1])
            self.assertEqual("Team Names", worksheet["A1"].value)
            self.assertEqual("a2", worksheet["A2"].value)
            self.assertEqual("b2", worksheet["A3"].value)

            statsWithTitles = year2.statSheet().preferredOrderWithTitle()
            for row, teamId in enumerate([team.id for team in year2.teams]):
                for col, statWithTitle in enumerate(statsWithTitles):
                    char = get_column_letter(col + 2)
                    if row == 1:
                        # check stat header
                        self.assertEqual(statWithTitle[0], worksheet[f"{char}{row}"].value)
                    # check stat value
                    # due to Excel rounding values, we assert that the values are very, very close
                    assert math.isclose(float(statWithTitle[1][teamId]), worksheet[f"{char}{row + 2}"].value,
                                        rel_tol=0.000000000000001)
